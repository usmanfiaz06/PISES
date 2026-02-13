#!/usr/bin/env python3
"""
PISES New Campus – Unit-Based Donor Pricing Spreadsheet Generator
Generates a professional Excel workbook for sharing with potential donors.

Based on:
  - Total BUA: ~52,400 m²
  - Mid-Institutional Cost: SAR 250,000,000 (midpoint of SAR 240-260M range)
  - Cost/m² BUA: ~SAR 4,771
  - Grossing factors: Academic 1.45×, High-Service 1.65×, Operations 1.55×
  - 1 USD = 3.75 SAR
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from copy import copy

# ── Constants ──────────────────────────────────────────────────────────────
TOTAL_COST_SAR = 250_000_000
TOTAL_BUA = 52_400
COST_PER_BUA_M2 = TOTAL_COST_SAR / TOTAL_BUA  # ~4,771
SAR_TO_USD = 1 / 3.75

# Grossing factors (NET → BUA)
GF_ACADEMIC = 1.45
GF_HIGH_SERVICE = 1.65
GF_OPERATIONS = 1.55

# ── Styles ─────────────────────────────────────────────────────────────────
DARK_GREEN = "1B5E20"
MED_GREEN = "388E3C"
LIGHT_GREEN = "E8F5E9"
ACCENT_GREEN = "C8E6C9"
WHITE = "FFFFFF"
GOLD = "F9A825"
LIGHT_GOLD = "FFF8E1"
DARK_GRAY = "333333"
MED_GRAY = "666666"
LIGHT_GRAY = "F5F5F5"
BORDER_COLOR = "BDBDBD"

thin_border = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)

header_font = Font(name="Calibri", bold=True, size=11, color=WHITE)
header_fill = PatternFill(start_color=DARK_GREEN, end_color=DARK_GREEN, fill_type="solid")
category_font = Font(name="Calibri", bold=True, size=11, color=DARK_GREEN)
category_fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
data_font = Font(name="Calibri", size=10, color=DARK_GRAY)
money_font = Font(name="Calibri", size=10, color=DARK_GRAY)
title_font = Font(name="Calibri", bold=True, size=22, color=DARK_GREEN)
subtitle_font = Font(name="Calibri", size=14, color=MED_GRAY)
section_font = Font(name="Calibri", bold=True, size=13, color=DARK_GREEN)
gold_fill = PatternFill(start_color=LIGHT_GOLD, end_color=LIGHT_GOLD, fill_type="solid")
alt_row_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")


def cost_per_unit(net_m2, grossing_factor):
    """Calculate construction cost for one unit given NET area and grossing factor."""
    bua = net_m2 * grossing_factor
    return round(bua * COST_PER_BUA_M2)


def usd(sar):
    return round(sar * SAR_TO_USD)


# ── Unit Data ──────────────────────────────────────────────────────────────
# Each tuple: (unit_name, description, qty, net_m2, grossing_factor, students_served_note)

UNITS = [
    # ── CLASSROOMS ─────────────────────────────────────────────────────
    ("CLASSROOMS & TEACHING SPACES", None, None, None, None, None),
    ("Standard Classroom (Grades 1–12)",
     "Fully equipped classroom with smart board, furniture & AC for 25 students",
     249, 43.12, GF_ACADEMIC, "25 students per classroom"),
    ("Kindergarten Classroom",
     "Purpose-built early-years classroom with play area, in-class washroom & learning corners",
     26, 62.5, GF_ACADEMIC, "25 children per classroom"),
    ("Nursery Activity Room",
     "Safe, stimulating activity space for youngest learners with age-appropriate furniture",
     9, 45.0, GF_ACADEMIC, "20–25 children per room"),
    ("Nursery Bedroom / Rest Room",
     "Dedicated rest area for nursery children with cots and soft furnishings",
     9, 22.5, GF_ACADEMIC, "20–25 children per room"),
    ("Reception Classroom",
     "Transition classroom bridging nursery to KG with learning stations",
     20, 62.5, GF_ACADEMIC, "25 children per classroom"),
    ("Early Years Learning Commons",
     "Shared indoor play and exploration zone for nursery & KG wings",
     3, 120.0, GF_ACADEMIC, "150–200 children per commons"),
    ("Primary Multi-Purpose Room",
     "Flexible teaching space for group work, presentations & project-based learning",
     4, 42.5, GF_ACADEMIC, "25–50 students per session"),

    # ── SCIENCE LABS ───────────────────────────────────────────────────
    ("SCIENCE LABORATORIES", None, None, None, None, None),
    ("Primary Science Lab",
     "Introductory science lab with demonstration bench, sinks & safety equipment",
     13, 60.1, GF_HIGH_SERVICE, "25 students per lab session"),
    ("Intermediate Science Lab",
     "General science lab with individual workstations, gas taps & fume extraction",
     7, 62.96, GF_HIGH_SERVICE, "25 students per lab session"),
    ("Secondary Science Lab (Physics / Chemistry / Biology)",
     "Specialist lab with discipline-specific equipment, data-logging & safety systems",
     12, 69.9, GF_HIGH_SERVICE, "25 students per lab session"),
    ("Science Prep Room",
     "Secure preparation and storage area serving a cluster of science labs",
     6, 18.0, GF_HIGH_SERVICE, "Supports 2–3 labs each"),
    ("Chemical Storage Room",
     "Ventilated, fire-rated chemical store with bunding & safety shower",
     2, 12.0, GF_HIGH_SERVICE, "Supports all secondary labs"),

    # ── ICT & COMPUTER LABS ────────────────────────────────────────────
    ("COMPUTER & ICT LABS", None, None, None, None, None),
    ("Primary Computer / Language Lab",
     "Age-appropriate computing lab with tablets/PCs and language learning software",
     6, 60.1, GF_HIGH_SERVICE, "25 students per session"),
    ("Secondary Computer / Language Lab",
     "Full ICT suite with desktop PCs, coding stations & language lab capability",
     10, 69.9, GF_HIGH_SERVICE, "25 students per session"),

    # ── SPECIALIST STUDIOS ─────────────────────────────────────────────
    ("SPECIALIST STUDIOS & MAKER SPACES", None, None, None, None, None),
    ("Maker / Robotics Lab",
     "Innovation hub with 3D printers, robotics kits, electronics workbenches",
     2, 120.0, GF_HIGH_SERVICE, "25 students per session"),
    ("Art Studio",
     "Creative space with easels, kiln access, wet & dry zones, natural lighting",
     2, 90.0, GF_HIGH_SERVICE, "25 students per session"),
    ("Primary Art Atelier",
     "Hands-on art workshop for younger students with washable materials & display areas",
     4, 41.9, GF_ACADEMIC, "25 students per session"),
    ("Music / Drama Room",
     "Acoustically treated performance & rehearsal space with instrument storage",
     2, 80.0, GF_HIGH_SERVICE, "30–40 students per session"),

    # ── LIBRARIES & LEARNING RESOURCE CENTRES ──────────────────────────
    ("LIBRARIES & LEARNING RESOURCE CENTRES", None, None, None, None, None),
    ("Primary Library / LRC",
     "Welcoming reading space with age-graded book collections & storytelling area",
     2, 75.1, GF_ACADEMIC, "40–60 students at a time"),
    ("Intermediate LRC",
     "Research-capable library with digital catalogue, reading nooks & group study",
     2, 74.7, GF_ACADEMIC, "40–60 students at a time"),
    ("Secondary LRC",
     "Advanced learning resource centre with digital research stations & quiet study",
     2, 88.6, GF_ACADEMIC, "50–70 students at a time"),

    # ── SPORTS & PHYSICAL EDUCATION ────────────────────────────────────
    ("SPORTS & PHYSICAL EDUCATION", None, None, None, None, None),
    ("Indoor Multi-Purpose Sports Hall",
     "Full-size covered hall for basketball, volleyball, badminton, futsal & events",
     2, 900.0, GF_HIGH_SERVICE, "200+ students per day"),
    ("25m Swimming Pool Complex",
     "6-lane pool with filtration plant, changing facilities, lifeguard station & spectator area",
     1, 1717.0, GF_HIGH_SERVICE, "300+ students per week"),
    ("Sports Changing & Shower Room",
     "Modern changing facility with lockers, showers & accessible cubicles",
     4, 160.0, GF_HIGH_SERVICE, "30–40 users per session"),
    ("Sports Storage Room",
     "Secure storage for PE equipment, balls, mats & sports gear",
     4, 25.0, GF_HIGH_SERVICE, "Supports all sports facilities"),
    ("Outdoor Multi-Sport Court",
     "Hard court for basketball, volleyball or tennis with line markings & lighting",
     6, 450.0, GF_HIGH_SERVICE, "30–40 students per court"),

    # ── DINING & FOOD SERVICES ─────────────────────────────────────────
    ("DINING & FOOD SERVICES", None, None, None, None, None),
    ("Dining Hall (700-seat, multi-shift)",
     "Full-service dining hall with fixed seating, servery counter & acoustic treatment",
     2, 1100.0, GF_HIGH_SERVICE, "700 students per sitting"),
    ("Commercial Kitchen & Prep Area",
     "Industrial kitchen with cooking stations, wash-up, cold rooms & dry stores",
     2, 300.0, GF_HIGH_SERVICE, "Serves 3,500 meals per day each"),
    ("Cold Room / Dry Store",
     "Temperature-controlled food storage with shelving and inventory management",
     4, 37.5, GF_OPERATIONS, "Supports kitchen operations"),

    # ── AUDITORIUM & ASSEMBLY ──────────────────────────────────────────
    ("AUDITORIUM & ASSEMBLY SPACES", None, None, None, None, None),
    ("Auditorium (300 seats)",
     "Tiered seating performance hall with stage, backstage, AV system & lighting rig",
     1, 740.0, GF_HIGH_SERVICE, "300-seat events, assemblies, graduations"),
    ("Atrium / Learning Commons",
     "Grand central gathering space for exhibitions, fairs, assemblies & informal learning",
     1, 2000.0, GF_HIGH_SERVICE, "2,000+ students for whole-school events"),
    ("Seminar Room",
     "Flexible meeting/teaching space for workshops, parent meetings & PD sessions",
     4, 45.0, GF_ACADEMIC, "20–30 attendees per room"),
    ("Breakout Room (Glass-walled)",
     "Small collaborative space for group work, tutoring & student projects",
     8, 25.0, GF_ACADEMIC, "6–10 students per room"),

    # ── EXAM CENTRE ────────────────────────────────────────────────────
    ("EXAM CENTRE", None, None, None, None, None),
    ("Exam Hall (300 candidates)",
     "Dedicated examination hall with individual desks, invigilator stations & CCTV",
     1, 750.0, GF_ACADEMIC, "300 candidates per session"),
    ("Candidate Holding Room",
     "Waiting area for students before exams with seating and bag storage",
     2, 60.0, GF_ACADEMIC, "150 students per room"),

    # ── SEN & WELLBEING ────────────────────────────────────────────────
    ("SEN & STUDENT WELLBEING", None, None, None, None, None),
    ("SEN Resource Room (Small Group)",
     "Specialist learning room for small-group interventions and differentiated support",
     10, 25.0, GF_ACADEMIC, "4–8 students per session"),
    ("1:1 Assessment Room",
     "Private room for individual assessments, educational psychologist evaluations",
     8, 12.0, GF_ACADEMIC, "1 student at a time"),
    ("Speech & Language Therapy Room",
     "Equipped therapy space for speech-language pathologists with AV tools",
     4, 16.0, GF_ACADEMIC, "1–3 students per session"),
    ("Occupational Therapy Room",
     "Sensory-motor therapy space with specialist equipment and observation area",
     2, 20.0, GF_ACADEMIC, "1–3 students per session"),
    ("Sensory Room",
     "Calming environment with sensory equipment for students with regulation needs",
     2, 24.0, GF_ACADEMIC, "1–4 students per session"),
    ("Counsellor Room",
     "Private space for student counselling, pastoral care & parent consultations",
     4, 12.0, GF_ACADEMIC, "1–2 students per session"),
    ("Medical Clinic / Nurse Room",
     "School clinic with examination bed, first-aid supplies & medication storage",
     2, 20.0, GF_ACADEMIC, "All students in wing"),
    ("Isolation / Rest Room (Medical)",
     "Short-stay room for unwell students awaiting parent collection",
     2, 10.0, GF_ACADEMIC, "1 student at a time"),

    # ── STAFF & PROFESSIONAL DEVELOPMENT ───────────────────────────────
    ("STAFF & PROFESSIONAL DEVELOPMENT", None, None, None, None, None),
    ("Staff Workroom (Distributed)",
     "Teacher planning & collaboration room with desks, printers & resources",
     12, 45.0, GF_ACADEMIC, "8–12 staff per workroom"),
    ("Staff Lounge",
     "Comfortable break room for staff with kitchen, seating & relaxation area",
     2, 90.0, GF_ACADEMIC, "40–60 staff per lounge"),
    ("Teacher Training / PD Room",
     "Professional development room with AV, flexible seating & workshop layout",
     2, 45.0, GF_ACADEMIC, "30–40 staff per session"),

    # ── ADMINISTRATION ─────────────────────────────────────────────────
    ("ADMINISTRATION & GOVERNANCE", None, None, None, None, None),
    ("Reception & Welcome Desk",
     "Visitor reception with waiting area, security check-in & information display",
     2, 25.0, GF_ACADEMIC, "Campus entrance"),
    ("Principal's Office",
     "Executive office for school principal with meeting area",
     1, 20.0, GF_ACADEMIC, "School leadership"),
    ("Admissions Office",
     "Parent-facing admissions suite for enrollment, interviews & documentation",
     2, 18.0, GF_ACADEMIC, "Handles 500+ applications/year"),
    ("Finance & Cashier Office",
     "Secure finance office with fee collection counter and record keeping",
     2, 40.0, GF_ACADEMIC, "All financial operations"),
    ("Board / SMC Meeting Room",
     "Formal boardroom for governance meetings with AV and conferencing",
     1, 30.0, GF_ACADEMIC, "12–20 board members"),
    ("School Store / Bookshop",
     "Retail space for books, stationery & school supplies",
     1, 80.0, GF_ACADEMIC, "Serves all students"),
    ("Uniform Shop",
     "Dedicated retail space for school uniform fittings and sales",
     1, 60.0, GF_ACADEMIC, "Serves all students"),

    # ── IT & SECURITY ──────────────────────────────────────────────────
    ("IT INFRASTRUCTURE & SECURITY", None, None, None, None, None),
    ("Server Room / MDF",
     "Climate-controlled data centre with rack servers, UPS & network backbone",
     2, 20.0, GF_OPERATIONS, "Entire campus IT"),
    ("Main Security Control Room (CCTV)",
     "24/7 security monitoring centre with CCTV screens, access control & fire panel",
     1, 20.0, GF_OPERATIONS, "Entire campus security"),
    ("IT Helpdesk Office",
     "Technical support hub for staff and student IT issues",
     2, 18.0, GF_OPERATIONS, "All campus users"),

    # ── TRANSPORT ──────────────────────────────────────────────────────
    ("TRANSPORT & LOGISTICS", None, None, None, None, None),
    ("Transport Office & Driver Lounge",
     "Bus fleet management office with driver rest area, lockers & washrooms",
     1, 43.0, GF_OPERATIONS, "70 buses, 2,400+ bus students"),

    # ── PRAYER & SPIRITUAL ─────────────────────────────────────────────
    ("PRAYER & SPIRITUAL SPACES", None, None, None, None, None),
    ("Prayer Room / Musalla",
     "Dedicated prayer space with ablution facilities, carpet & qibla marker",
     4, 60.0, GF_ACADEMIC, "100–150 worshippers per room"),
]

# ── Donor Package Tiers ────────────────────────────────────────────────────
PACKAGES = [
    ("INDIVIDUAL IMPACT GIFTS", "SAR 50K – 500K", [
        ("Name a Classroom", "Sponsor one standard classroom with naming recognition",
         "Standard Classroom (Grades 1–12)", 1),
        ("Equip a Science Lab", "Fund one fully-equipped primary science lab",
         "Primary Science Lab", 1),
        ("Build a Sensory Room", "Provide a calming sensory room for SEN students",
         "Sensory Room", 1),
        ("Sponsor a Library Corner", "Fund one primary library / LRC",
         "Primary Library / LRC", 1),
        ("Create an Art Atelier", "Build one primary art workshop for budding artists",
         "Primary Art Atelier", 1),
    ]),
    ("MAJOR GIFTS", "SAR 500K – 5M", [
        ("Robotics Innovation Hub", "Fund one maker/robotics lab for STEM education",
         "Maker / Robotics Lab", 1),
        ("Auditorium Naming", "Sponsor the 300-seat auditorium for school events",
         "Auditorium (300 seats)", 1),
        ("Sports Hall Sponsor", "Fund one indoor sports hall for 200+ students/day",
         "Indoor Multi-Purpose Sports Hall", 1),
        ("Dining Experience", "Sponsor one dining hall serving 700 students per sitting",
         "Dining Hall (700-seat, multi-shift)", 1),
        ("Classroom Block (10 rooms)", "Build a block of 10 classrooms for 250 students",
         "Standard Classroom (Grades 1–12)", 10),
    ]),
    ("LANDMARK GIFTS", "SAR 5M+", [
        ("Swimming Pool Complex", "Fund the entire 25m pool with all support facilities",
         "25m Swimming Pool Complex", 1),
        ("Exam Centre", "Sponsor the 300-candidate exam hall with holding rooms",
         None, None),  # Custom calculation
        ("Learning Commons & Atrium", "Fund the grand 2,000 m² central gathering space",
         "Atrium / Learning Commons", 1),
        ("Entire Early Years Wing", "Build all nursery & KG classrooms (55 rooms)",
         None, None),  # Custom
        ("Complete SEN Suite", "Fund the entire SEN & wellbeing department (34 rooms)",
         None, None),  # Custom
    ]),
]


def build_workbook():
    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 1: UNIT PRICING
    # ══════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Unit Pricing"
    ws.sheet_properties.tabColor = DARK_GREEN

    # Column widths
    col_widths = [4, 42, 58, 8, 10, 12, 18, 18, 14, 14, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title block
    ws.merge_cells("B1:K1")
    ws["B1"] = "PISES NEW CAMPUS — UNIT-BASED DONOR PRICING"
    ws["B1"].font = title_font
    ws["B1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("B2:K2")
    ws["B2"] = "Pakistan International School (English Section), Riyadh  |  7,000-Student Campus  |  SAR 250 Million Project  |  Prices in 2025 SAR"
    ws["B2"].font = subtitle_font
    ws["B2"].alignment = Alignment(vertical="center")
    ws.row_dimensions[2].height = 25

    ws.merge_cells("B3:K3")
    ws["B3"] = "1 USD = 3.75 SAR  |  Prices include construction, MEP, fit-out, ICT & furniture  |  Excluding land, professional fees & inflation"
    ws["B3"].font = Font(name="Calibri", size=10, italic=True, color=MED_GRAY)
    ws.row_dimensions[3].height = 20

    # Headers (row 5)
    headers = [
        "#", "Unit Name", "Description", "Qty",
        "NET m²", "BUA m²", "Cost / Unit (SAR)", "Cost / Unit (USD)",
        "Total (SAR)", "Total (USD)", "Students Impacted"
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[5].height = 30

    row = 6
    item_num = 0
    unit_lookup = {}  # For package sheet cross-reference
    grand_total_sar = 0

    for entry in UNITS:
        name, desc, qty, net_m2, gf, students = entry

        if qty is None:
            # Category header row
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
            cell = ws.cell(row=row, column=1, value=name)
            cell.font = category_font
            cell.fill = category_fill
            cell.alignment = Alignment(vertical="center")
            cell.border = thin_border
            for c in range(2, 12):
                ws.cell(row=row, column=c).fill = category_fill
                ws.cell(row=row, column=c).border = thin_border
            ws.row_dimensions[row].height = 28
            row += 1
            continue

        item_num += 1
        unit_cost_sar = cost_per_unit(net_m2, gf)
        bua_m2 = round(net_m2 * gf, 1)
        total_sar = unit_cost_sar * qty
        grand_total_sar += total_sar

        unit_lookup[name] = {
            "unit_cost_sar": unit_cost_sar,
            "net_m2": net_m2,
            "qty": qty,
        }

        values = [
            item_num, name, desc, qty,
            round(net_m2, 1), bua_m2,
            unit_cost_sar, usd(unit_cost_sar),
            total_sar, usd(total_sar),
            students
        ]

        is_alt = (item_num % 2 == 0)
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            if is_alt:
                cell.fill = alt_row_fill

            # Alignment & formatting
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in (2, 3, 11):
                cell.alignment = Alignment(wrap_text=True, vertical="center")
            elif col_idx == 4:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in (5, 6):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,##0.0"
            elif col_idx in (7, 8, 9, 10):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,##0"

        ws.row_dimensions[row].height = 36
        row += 1

    # Grand total row
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=1, value="GRAND TOTAL (All Units)")
    cell.font = Font(name="Calibri", bold=True, size=12, color=WHITE)
    cell.fill = PatternFill(start_color=DARK_GREEN, end_color=DARK_GREEN, fill_type="solid")
    cell.alignment = Alignment(horizontal="right", vertical="center")
    for c in range(1, 12):
        ws.cell(row=row, column=c).fill = PatternFill(start_color=DARK_GREEN, end_color=DARK_GREEN, fill_type="solid")
        ws.cell(row=row, column=c).border = thin_border
        ws.cell(row=row, column=c).font = Font(name="Calibri", bold=True, size=12, color=WHITE)

    ws.cell(row=row, column=9, value=grand_total_sar).number_format = "#,##0"
    ws.cell(row=row, column=9).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=row, column=10, value=usd(grand_total_sar)).number_format = "#,##0"
    ws.cell(row=row, column=10).alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[row].height = 30

    # Note rows
    row += 2
    notes = [
        "NOTES:",
        "1. All prices are planning-level estimates based on mid-institutional specification (2025 SAR baseline).",
        "2. Prices include: construction, structural, MEP (mechanical/electrical/plumbing), interior fit-out, ICT infrastructure, furniture & equipment.",
        "3. Prices EXCLUDE: land cost, architectural/engineering professional fees, financing costs, inflation beyond 2025.",
        "4. Naming rights and recognition plaques available for donors of individual units.",
        "5. Donor contributions are cumulative — multiple donors may co-sponsor larger facilities.",
        "6. All facilities comply with Saudi Building Code 2024 and TBC Category A standards.",
        f"7. Grand total reflects sum of all individual units. Full campus cost: SAR 240–260 Million (mid-range: SAR 250M).",
    ]
    for note in notes:
        cell = ws.cell(row=row, column=2, value=note)
        if note == "NOTES:":
            cell.font = Font(name="Calibri", bold=True, size=10, color=DARK_GREEN)
        else:
            cell.font = Font(name="Calibri", size=9, color=MED_GRAY, italic=True)
        row += 1

    # Freeze panes
    ws.freeze_panes = "A6"

    # Print setup
    ws.print_area = f"A1:K{row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.fitToWidth = 1

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 2: DONOR PACKAGES
    # ══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Donor Packages")
    ws2.sheet_properties.tabColor = GOLD

    col_widths2 = [4, 35, 55, 20, 20, 22]
    for i, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws2.merge_cells("B1:F1")
    ws2["B1"] = "PISES NEW CAMPUS — DONOR PACKAGES"
    ws2["B1"].font = title_font
    ws2.row_dimensions[1].height = 40

    ws2.merge_cells("B2:F2")
    ws2["B2"] = "Suggested giving levels with naming recognition  |  All amounts in SAR & USD"
    ws2["B2"].font = subtitle_font
    ws2.row_dimensions[2].height = 25

    # Custom cost calculations for special packages
    exam_centre_cost = (cost_per_unit(750, GF_ACADEMIC) +
                        2 * cost_per_unit(60, GF_ACADEMIC) +
                        2 * cost_per_unit(25, GF_ACADEMIC))
    ey_wing_cost = (9 * cost_per_unit(45, GF_ACADEMIC) +
                    9 * cost_per_unit(22.5, GF_ACADEMIC) +
                    20 * cost_per_unit(62.5, GF_ACADEMIC) +
                    26 * cost_per_unit(62.5, GF_ACADEMIC) +
                    3 * cost_per_unit(120, GF_ACADEMIC))
    sen_suite_cost = (10 * cost_per_unit(25, GF_ACADEMIC) +
                      8 * cost_per_unit(12, GF_ACADEMIC) +
                      4 * cost_per_unit(16, GF_ACADEMIC) +
                      2 * cost_per_unit(20, GF_ACADEMIC) +
                      2 * cost_per_unit(24, GF_ACADEMIC) +
                      4 * cost_per_unit(12, GF_ACADEMIC) +
                      2 * cost_per_unit(20, GF_ACADEMIC) +
                      2 * cost_per_unit(10, GF_ACADEMIC))

    pkg_row = 4
    for tier_name, tier_range, items in PACKAGES:
        # Tier header
        ws2.merge_cells(start_row=pkg_row, start_column=1, end_row=pkg_row, end_column=6)
        cell = ws2.cell(row=pkg_row, column=1, value=f"{tier_name}  ({tier_range})")
        cell.font = Font(name="Calibri", bold=True, size=13, color=WHITE)
        cell.fill = PatternFill(start_color=MED_GREEN, end_color=MED_GREEN, fill_type="solid")
        cell.alignment = Alignment(vertical="center")
        for c in range(1, 7):
            ws2.cell(row=pkg_row, column=c).fill = PatternFill(start_color=MED_GREEN, end_color=MED_GREEN, fill_type="solid")
            ws2.cell(row=pkg_row, column=c).border = thin_border
        ws2.row_dimensions[pkg_row].height = 32
        pkg_row += 1

        # Sub-headers
        sub_headers = ["#", "Package Name", "What You Fund", "Amount (SAR)", "Amount (USD)", "Impact"]
        for col_idx, h in enumerate(sub_headers, 1):
            cell = ws2.cell(row=pkg_row, column=col_idx, value=h)
            cell.font = Font(name="Calibri", bold=True, size=10, color=DARK_GREEN)
            cell.fill = PatternFill(start_color=ACCENT_GREEN, end_color=ACCENT_GREEN, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws2.row_dimensions[pkg_row].height = 24
        pkg_row += 1

        for idx, (pkg_name, pkg_desc, ref_unit, ref_qty) in enumerate(items, 1):
            # Calculate cost
            if ref_unit and ref_unit in unit_lookup:
                pkg_cost = unit_lookup[ref_unit]["unit_cost_sar"] * ref_qty
                impact = f"{ref_qty * 25} students" if ref_qty > 1 else f"25 students"
            elif pkg_name == "Exam Centre":
                pkg_cost = exam_centre_cost
                impact = "300 candidates/session"
            elif pkg_name == "Entire Early Years Wing":
                pkg_cost = ey_wing_cost
                impact = "800+ young learners"
            elif pkg_name == "Complete SEN Suite":
                pkg_cost = sen_suite_cost
                impact = "500+ students with special needs"
            else:
                pkg_cost = 0
                impact = ""

            # Special impact notes
            if "Auditorium" in pkg_name:
                impact = "300-seat events & graduations"
            elif "Sports Hall" in pkg_name:
                impact = "200+ students/day"
            elif "Dining" in pkg_name:
                impact = "700 students/sitting"
            elif "Swimming" in pkg_name:
                impact = "300+ students/week"
            elif "Learning Commons" in pkg_name:
                impact = "2,000+ for whole-school events"
            elif "Robotics" in pkg_name:
                impact = "STEM for 25 students/session"
            elif "Sensory" in pkg_name:
                impact = "Students with regulation needs"
            elif "Library" in pkg_name:
                impact = "40–60 students at a time"
            elif "Art Atelier" in pkg_name:
                impact = "25 young artists/session"

            values = [idx, pkg_name, pkg_desc, pkg_cost, usd(pkg_cost), impact]
            is_alt = (idx % 2 == 0)
            for col_idx, val in enumerate(values, 1):
                cell = ws2.cell(row=pkg_row, column=col_idx, value=val)
                cell.font = data_font
                cell.border = thin_border
                if is_alt:
                    cell.fill = alt_row_fill
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_idx in (4, 5):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = "#,##0"
                else:
                    cell.alignment = Alignment(wrap_text=True, vertical="center")
            ws2.row_dimensions[pkg_row].height = 32
            pkg_row += 1

        pkg_row += 1  # Spacing between tiers

    # Package notes
    pkg_row += 1
    pkg_notes = [
        "HOW TO GIVE:",
        "• Donors may sponsor any unit individually or combine units for larger impact.",
        "• Naming rights available for gifts of SAR 250,000 and above (recognition plaque on facility).",
        "• Co-sponsorship welcomed — multiple donors can share the cost of larger facilities.",
        "• All donations are tax-deductible where applicable under local regulations.",
        "• Contact the PISES Development Office for customized giving plans and recognition.",
    ]
    for note in pkg_notes:
        cell = ws2.cell(row=pkg_row, column=2, value=note)
        if note == "HOW TO GIVE:":
            cell.font = Font(name="Calibri", bold=True, size=11, color=DARK_GREEN)
        else:
            cell.font = Font(name="Calibri", size=10, color=MED_GRAY)
        pkg_row += 1

    ws2.freeze_panes = "A4"

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 3: SUMMARY BY CATEGORY
    # ══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Category Summary")
    ws3.sheet_properties.tabColor = "1565C0"

    col_widths3 = [4, 40, 10, 16, 20, 20, 12]
    for i, w in enumerate(col_widths3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    ws3.merge_cells("B1:G1")
    ws3["B1"] = "PISES NEW CAMPUS — COST SUMMARY BY CATEGORY"
    ws3["B1"].font = title_font
    ws3.row_dimensions[1].height = 40

    ws3.merge_cells("B2:G2")
    ws3["B2"] = "High-level overview for donor briefings  |  7,000-Student Campus"
    ws3["B2"].font = subtitle_font
    ws3.row_dimensions[2].height = 25

    # Build category totals from UNITS data
    category_totals = []
    current_cat = None
    cat_units = 0
    cat_net = 0.0
    cat_cost = 0

    for entry in UNITS:
        name, desc, qty, net_m2, gf, students = entry
        if qty is None:
            # Save previous category
            if current_cat:
                category_totals.append((current_cat, cat_units, cat_net, cat_cost))
            current_cat = name
            cat_units = 0
            cat_net = 0.0
            cat_cost = 0
        else:
            cat_units += qty
            cat_net += net_m2 * qty
            cat_cost += cost_per_unit(net_m2, gf) * qty

    # Don't forget last category
    if current_cat:
        category_totals.append((current_cat, cat_units, cat_net, cat_cost))

    # Headers
    sum_headers = ["#", "Category", "Units", "Total NET m²", "Total Cost (SAR)", "Total Cost (USD)", "% of Budget"]
    for col_idx, h in enumerate(sum_headers, 1):
        cell = ws3.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws3.row_dimensions[4].height = 28

    sum_row = 5
    overall_total = sum(c[3] for c in category_totals)

    for idx, (cat_name, units, net, cost) in enumerate(category_totals, 1):
        pct = cost / overall_total * 100 if overall_total else 0
        values = [idx, cat_name, units, round(net, 0), cost, usd(cost), round(pct, 1)]
        is_alt = (idx % 2 == 0)
        for col_idx, val in enumerate(values, 1):
            cell = ws3.cell(row=sum_row, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            if is_alt:
                cell.fill = alt_row_fill
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 2:
                cell.alignment = Alignment(wrap_text=True, vertical="center")
            elif col_idx == 3:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in (4, 5, 6):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,##0"
            elif col_idx == 7:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = "0.0\"%\""
        ws3.row_dimensions[sum_row].height = 28
        sum_row += 1

    # Grand total
    sum_row += 1
    ws3.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=3)
    for c in range(1, 8):
        ws3.cell(row=sum_row, column=c).fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        ws3.cell(row=sum_row, column=c).font = Font(name="Calibri", bold=True, size=11, color=WHITE)
        ws3.cell(row=sum_row, column=c).border = thin_border
    ws3.cell(row=sum_row, column=1, value="GRAND TOTAL").alignment = Alignment(horizontal="right", vertical="center")
    ws3.cell(row=sum_row, column=4, value=sum(c[2] for c in category_totals)).number_format = "#,##0"
    ws3.cell(row=sum_row, column=4).alignment = Alignment(horizontal="right", vertical="center")
    ws3.cell(row=sum_row, column=5, value=overall_total).number_format = "#,##0"
    ws3.cell(row=sum_row, column=5).alignment = Alignment(horizontal="right", vertical="center")
    ws3.cell(row=sum_row, column=6, value=usd(overall_total)).number_format = "#,##0"
    ws3.cell(row=sum_row, column=6).alignment = Alignment(horizontal="right", vertical="center")
    ws3.cell(row=sum_row, column=7, value="100%").alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[sum_row].height = 30

    ws3.freeze_panes = "A5"

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 4: QUICK REFERENCE (Single page for donors)
    # ══════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Quick Reference")
    ws4.sheet_properties.tabColor = GOLD

    col_widths4 = [4, 38, 18, 18]
    for i, w in enumerate(col_widths4, 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    ws4.merge_cells("B1:D1")
    ws4["B1"] = "WHAT YOUR GIFT CAN BUILD"
    ws4["B1"].font = Font(name="Calibri", bold=True, size=24, color=DARK_GREEN)
    ws4.row_dimensions[1].height = 45

    ws4.merge_cells("B2:D2")
    ws4["B2"] = "PISES New Campus  |  Every contribution builds a future"
    ws4["B2"].font = Font(name="Calibri", size=13, color=MED_GRAY)
    ws4.row_dimensions[2].height = 28

    # Quick reference items — most donor-friendly picks
    quick_items = [
        ("SAR 50,000 – 100,000", None),
        ("SEN Assessment Room", "SAR 83,000 / USD 22,000"),
        ("Counsellor Room", "SAR 83,000 / USD 22,000"),
        ("Breakout Room", "SAR 173,000 / USD 46,000"),
        ("Medical Clinic", "SAR 138,000 / USD 37,000"),
        ("", ""),
        ("SAR 100,000 – 300,000", None),
        ("Nursery Bedroom", "SAR 156,000 / USD 42,000"),
        ("SEN Resource Room", "SAR 173,000 / USD 46,000"),
        ("Primary Art Atelier", "SAR 290,000 / USD 77,000"),
        ("Standard Classroom", "SAR 298,000 / USD 79,000"),
        ("", ""),
        ("SAR 300,000 – 500,000", None),
        ("Nursery Activity Room", "SAR 311,000 / USD 83,000"),
        ("KG / Reception Classroom", "SAR 432,000 / USD 115,000"),
        ("Primary Science Lab", "SAR 473,000 / USD 126,000"),
        ("Primary Computer Lab", "SAR 473,000 / USD 126,000"),
        ("", ""),
        ("SAR 500,000 – 1,000,000", None),
        ("Secondary Science Lab", "SAR 551,000 / USD 147,000"),
        ("Secondary Computer Lab", "SAR 551,000 / USD 147,000"),
        ("Music / Drama Room", "SAR 630,000 / USD 168,000"),
        ("Art Studio", "SAR 709,000 / USD 189,000"),
        ("Early Years Learning Commons", "SAR 830,000 / USD 221,000"),
        ("Maker / Robotics Lab", "SAR 945,000 / USD 252,000"),
        ("", ""),
        ("SAR 1,000,000 – 5,000,000", None),
        ("Prayer Room / Musalla", "SAR 415,000 / USD 111,000 (×4 = SAR 1.66M)"),
        ("Classroom Block (10 rooms)", "SAR 2,980,000 / USD 795,000"),
        ("Exam Hall (300 candidates)", "SAR 5,188,000 / USD 1,383,000"),
        ("", ""),
        ("SAR 5,000,000+", None),
        ("Indoor Sports Hall", "SAR 7,085,000 / USD 1,889,000"),
        ("Dining Hall + Kitchen", "SAR 11,020,000 / USD 2,939,000"),
        ("Swimming Pool Complex", "SAR 13,519,000 / USD 3,605,000"),
        ("Auditorium (300 seats)", "SAR 5,824,000 / USD 1,553,000"),
        ("Atrium / Learning Commons", "SAR 15,741,000 / USD 4,198,000"),
    ]

    qr_row = 4
    for item_name, item_cost in quick_items:
        if item_name == "":
            qr_row += 1
            continue
        if item_cost is None:
            # Tier header
            ws4.merge_cells(start_row=qr_row, start_column=2, end_row=qr_row, end_column=4)
            cell = ws4.cell(row=qr_row, column=2, value=item_name)
            cell.font = Font(name="Calibri", bold=True, size=12, color=WHITE)
            cell.fill = PatternFill(start_color=MED_GREEN, end_color=MED_GREEN, fill_type="solid")
            cell.alignment = Alignment(vertical="center")
            for c in range(2, 5):
                ws4.cell(row=qr_row, column=c).fill = PatternFill(start_color=MED_GREEN, end_color=MED_GREEN, fill_type="solid")
                ws4.cell(row=qr_row, column=c).border = thin_border
            ws4.row_dimensions[qr_row].height = 28
        else:
            ws4.cell(row=qr_row, column=2, value=item_name).font = Font(name="Calibri", size=11, color=DARK_GRAY)
            ws4.cell(row=qr_row, column=2).border = thin_border
            ws4.merge_cells(start_row=qr_row, start_column=3, end_row=qr_row, end_column=4)
            ws4.cell(row=qr_row, column=3, value=item_cost).font = Font(name="Calibri", size=11, bold=True, color=DARK_GREEN)
            ws4.cell(row=qr_row, column=3).alignment = Alignment(horizontal="right", vertical="center")
            ws4.cell(row=qr_row, column=3).border = thin_border
            for c in range(3, 5):
                ws4.cell(row=qr_row, column=c).border = thin_border
            ws4.row_dimensions[qr_row].height = 24
        qr_row += 1

    ws4.freeze_panes = "A4"

    # ── Save ───────────────────────────────────────────────────────────
    output_path = "/home/user/PISES/PISES_Donor_Unit_Pricing.xlsx"
    wb.save(output_path)
    print(f"✓ Workbook saved: {output_path}")
    print(f"  Sheets: {wb.sheetnames}")
    print(f"  Grand total (all units): SAR {grand_total_sar:,.0f} / USD {usd(grand_total_sar):,.0f}")
    print(f"  Category summary total: SAR {overall_total:,.0f}")
    return output_path


if __name__ == "__main__":
    build_workbook()
